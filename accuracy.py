import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from config import (

  GEMINI_TEST,OPENAI_TEST
)
# Load the Excel file
file_path = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\accuracy_report.xlsx'
xls = pd.ExcelFile(file_path)

# Read the sheets into DataFrames
sheet1 = pd.read_excel(xls, sheet_name='Sheet1')
sheet2 = pd.read_excel(xls, sheet_name='Sheet2')
print(sheet1.shape)

def calculate_scores_and_out_of_range(sheet, column_index, min_value, max_value, start_row):
    # Get values in the specified column starting from start_row (zero-based indexing)
    column_values = sheet.iloc[start_row:, column_index]

    # Convert to float and filter out non-numeric values
    filtered_values = pd.to_numeric(column_values, errors='coerce').dropna()

    # Calculate the number of values between min_value and max_value
    count_in_range = ((filtered_values >= min_value) & (filtered_values <= max_value)).sum()

    # Identify row numbers with values out of the specified range
    out_of_range_rows = column_values[(column_values < min_value) | (column_values > max_value)].index + 2  # +1 to match Excel row indexing

    # Calculate the total number of values
    total_values = filtered_values.size

    # Calculate the ratio
    ratio = count_in_range / total_values if total_values != 0 else 0

    return count_in_range, out_of_range_rows, ratio

def main():

    if GEMINI_TEST==True:
        col1=1
        col2=3
        col3=5

    if OPENAI_TEST==True:
        col1=7
        col2=9
        col3=11
    columns_and_ranges = [
        (col1, 7, 10),  # Column 2 (Excel-based), values between 7 and 10
        (col2, 4, 7),   # Column 4 (Excel-based), values between 4 and 7
        (col3, 0, 4)    # Column 6 (Excel-based), values between 0 and 4
    ]

    update_data = {}
    ratios = []

    # Load the workbook and select the sheet to update
    wb = load_workbook(file_path)
    ws = wb['Sheet1']

    # Find the next available row in Sheet1
    next_row = ws.max_row+1

    for i, (col_index, min_val, max_val) in enumerate(columns_and_ranges):
        count_in_range, out_of_range_rows, ratio = calculate_scores_and_out_of_range(sheet2, col_index, min_val, max_val, 3)
        print(f'Column {col_index + 1} (Values between {min_val} and {max_val}):')
        print(f'  Number of scores in range: {count_in_range}')
        print(f'  Rows with values out of range: {out_of_range_rows.tolist()}')
        print(f'  Ratio: {ratio}')

        # Determine the correct columns to update based on iteration
        ratio_column = 3 + (i * 2)
        out_of_range_column = 4 + (i * 2)

        # Update the dictionary for each set of results
        update_data[(next_row, ratio_column)] = ratio  # Next available row, ratio column
        update_data[(next_row, out_of_range_column)] = ','.join(map(str, out_of_range_rows.tolist()))  # Next available row, out-of-range rows column

        # Collect the ratio for average calculation
        ratios.append(ratio)

    # Calculate the average ratio
    average_ratio = np.mean(ratios)

    # Update cells in Sheet1
    for (row, col), value in update_data.items():
        cell = ws.cell(row=row, column=col)  # No need to adjust since next_row is already correct
        cell.value = value

    # Add the current timestamp to column 2, next available row
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.cell(row=next_row, column=2).value = timestamp

    # Insert the average ratio into the 9th column of the next available row
    ws.cell(row=next_row, column=9).value = average_ratio

    # Save the updated workbook
    wb.save(file_path)

if __name__ == "__main__":
    main()
