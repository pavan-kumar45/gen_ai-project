import pandas as pd
import time
start_time=time.time()
import re
from google.api_core.exceptions import ResourceExhausted, InvalidArgument
from langchain_google_vertexai import ChatVertexAI
from langchain_core.messages import HumanMessage
import vertexai
from config import (
    SERVICE_ACCOUNT_FILE_PATH, PROJECT_ID, LOCATION,
    MODEL_PARAMS,Run_specific_row,RUN_SPECIFIC_ROW_INDEX,ACCURACY_TEST,REPEATABILITY_TEST
)
from prompts import theory_prompt_template, coding_prompt_template
from openpyxl import load_workbook
import subprocess
# Set up the Vertex AI environment
vertexai.init(project=PROJECT_ID, location=LOCATION)

# Initialize the Vertex AI model using LangChain's ChatVertexAI
llm = ChatVertexAI(**MODEL_PARAMS)

# Function to generate feedback with retry mechanism using ChatVertexAI
def generate_feedback(prompt, retries=6, delay=1, max_wait_time=300):
    total_wait_time = 0
    for attempt in range(retries):
        if total_wait_time >= max_wait_time:
            print("Total wait time exceeded. Stopping retries.")
            return None

        try:
            message = HumanMessage(content=prompt)
            response = llm.invoke([message])
            return response.content
        except (ResourceExhausted, InvalidArgument) as e:
            if isinstance(e, InvalidArgument) and "topK" in str(e):
                print("Invalid topK value. Please update the value to be within the supported range.")
                return None
            if attempt < retries - 1:
                backoff_time = delay * (2 ** attempt)  # Exponential backoff
                total_wait_time += backoff_time
                print(f"Quota exceeded. Retrying in {backoff_time} seconds...")
                time.sleep(backoff_time)
            else:
                print("Max retries exceeded. Please check your quota.")
                return None

# Function to extract score from feedback
def extract_score(feedback):
    match = re.search(r'\b\d{1,2}\b', feedback)
    return int(match.group()) if match else None

# Read the input Excel file
file_path = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\Question Bank.xlsx'
sheet_name = 'Sheet1'
column_pairs = [(0, 1), (0, 2), (0, 3)]

# Define the output Excel file
output_file_path = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\accuracy_report.xlsx'
output_sheet_name = 'Sheet2'  # Specify the sheet name to update

# Open the output workbook and select the desired sheet
wb = load_workbook(output_file_path)
ws = wb[output_sheet_name]

def run_all_rows(flag1):
    if not flag1:
        return
    
    for col_index, col_pair in enumerate(column_pairs):
        df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=list(col_pair))
        print(f"\nReading columns {col_pair[0]+1} and {col_pair[1]+1}:\n")

        # Determine the starting columns for scores and feedback in the output file
        score_col = 2 * col_index + 2
        feedback_col = score_col + 1

        # Iterate through each row of the dataframe
        for index, row in df.iterrows():
            question = row.iloc[0]
            answer = row.iloc[1]

            # Determine the type of question and choose the appropriate template
            if "program" in question.lower():
                prompt_template = coding_prompt_template
            else:
                prompt_template = theory_prompt_template

            prompt = prompt_template.format(question=question, user_answer=answer)

            feedback = generate_feedback(prompt)
            if feedback:
                score = extract_score(feedback)

                # Update the output Excel file directly in the specified sheet
                ws.cell(row=index + 4, column=score_col, value=score)  # Assuming row 4 for scores
                ws.cell(row=index + 4, column=feedback_col, value=feedback)  # Assuming row 4 for feedback

                print(f"Question: {question}")
                print(f"Score: {score}")
                print(f"Feedback: {feedback}\n")

                print("-----------------------------------------------------------------------------------")
                
                # Save the workbook after each update to reflect changes immediately
                wb.save(output_file_path)

def run_specific_row(row_index, col_pair_index, flag2):
    if not flag2:
        return
    
    col_pair = column_pairs[col_pair_index]
    df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=list(col_pair))
    print(f"\nReading columns {col_pair[0]+1} and {col_pair[1]+1} for row {row_index}:\n")

    # Determine the starting columns for scores and feedback in the output file
    score_col = 2 * col_pair_index + 2
    feedback_col = score_col + 1

    # Get the specific row
    row = df.iloc[row_index]

    question = row.iloc[0]
    answer = row.iloc[1]

    # Determine the type of question and choose the appropriate template
    if "program" in question.lower():
        prompt_template = coding_prompt_template
    else:
        prompt_template = theory_prompt_template

    prompt = prompt_template.format(question=question, user_answer=answer)

    feedback = generate_feedback(prompt)
    if feedback:
        score = extract_score(feedback)

        # Update the output Excel file directly in the specified sheet
        ws.cell(row=row_index + 4, column=score_col, value=score)  # Assuming row 4 for scores
        ws.cell(row=row_index + 4, column=feedback_col, value=feedback)  # Assuming row 4 for feedback

        print(f"Question: {question}")
        print(f"Score: {score}")
        print(f"Feedback: {feedback}\n")

        print("-----------------------------------------------------------------------------------")
        
        # Save the workbook after each update to reflect changes immediately
        wb.save(output_file_path)

# Example usage:
run_all_rows(flag1=ACCURACY_TEST)
actual_index=RUN_SPECIFIC_ROW_INDEX
run_specific_row(row_index=actual_index-4, col_pair_index=2, flag2=Run_specific_row)
if ACCURACY_TEST==True or Run_specific_row==True:
    subprocess.run(['python', 'accuracy.py'])
# Close the workbook after all updates are done
wb.close()

if REPEATABILITY_TEST==True:
    subprocess.run(['python', 'repeatibility.py'])
    subprocess.run(['python', 'repeatablity_report.py'])

elapsed_time = time.time() - start_time
print(f"Total time taken: {elapsed_time/60} minutes")
