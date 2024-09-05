import time
import openpyxl
from openpyxl import load_workbook
import re
from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage
from config import user_answer_written
# Import templates
from prompts_evaluate import theory_prompt_template, coding_prompt_template, design_prompt_template

# Initialize the ChatGroq model with the specified parameters
llm = ChatGroq(
    temperature=0,
    groq_api_key="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",  # Replace with your actual Groq API key
    model_name="llama-3.1-70b-versatile",
    max_tokens=1000,
    timeout=60
)

def generate_feedback(prompt_template, question, user_answer, retries=6, delay=1, max_wait_time=300):
    total_wait_time = 0
    prompt = prompt_template.format(question=question, user_answer=user_answer)
    for attempt in range(retries):
        if total_wait_time >= max_wait_time:
            print("Total wait time exceeded. Stopping retries.")
            return None

        try:
            message = HumanMessage(content=prompt)
            response = llm.invoke(message.content)  # Adjusted to use ChatGroq's invoke method
            return response.content
        except Exception as e:  # Using a generic exception handler
            print(f"Exception occurred: {e}")
            if attempt < retries - 1:
                backoff_time = delay * (2 ** attempt)  # Exponential backoff
                total_wait_time += backoff_time
                print(f"Retrying in {backoff_time} seconds...")
                time.sleep(backoff_time)
            else:
                print("Max retries exceeded. Please check your settings.")
                return None

def extract_score(feedback):
    match = re.search(r'\b\d{1,2}\b', feedback)
    return int(match.group()) if match else None

def process_excel(file_path):
    i = 1
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=False):  # Assuming the first row is the header
        question_cell, topic_cell, answer_cell, score_cell, feedback_cell = row

        question = question_cell.value
        topic = topic_cell.value
        user_answer = answer_cell.value

        if question is None or user_answer is None or topic is None:
            continue

        # Determine the prompt template based on the topic
        if topic.startswith("Theory"):
            prompt_template = theory_prompt_template
        elif topic.startswith("Coding"):
            prompt_template = coding_prompt_template
        elif topic.startswith("Design"):
            prompt_template = design_prompt_template
        else:
            print(f"Unrecognized topic: {topic}. Skipping row.")
            continue

        print(f"Evaluating Question: {i}")
        i += 1
        feedback = generate_feedback(prompt_template, question, user_answer)
        if feedback:
            score = extract_score(feedback)
            score_cell.value = score if score is not None else "N/A"
            feedback_cell.value = feedback

    workbook.save(file_path)
    print(f"Processed {file_path} and saved results.")

if user_answer_written:
    process_excel(r'C:\Users\pavan\Desktop\projects\llama_3.1\CSV Files\Questions.xlsx')
