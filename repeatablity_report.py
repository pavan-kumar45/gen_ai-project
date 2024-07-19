import openpyxl
import time
import ast
import numpy as np

start_time = time.time()

def combine_scores(input_file):
    # Load the input workbook
    wb = openpyxl.load_workbook(input_file)
    
    # Get the sheets
    sheets = [wb[f'Sheet{i}'] for i in range(2, 7)]
    
    # Check if 'Sheet1' exists, if not, create it
    if 'Sheet1' in wb.sheetnames:
        output_sheet = wb['Sheet1']
    else:
        output_sheet = wb.create_sheet('Sheet1')
    
    # Combine the scores
    for row_idx in range(4, sheets[0].max_row + 1):
        combined_row = [[], [], []]
        for sheet in sheets:
            combined_row[0].append(sheet.cell(row=row_idx, column=2).value)
            combined_row[1].append(sheet.cell(row=row_idx, column=4).value)
            combined_row[2].append(sheet.cell(row=row_idx, column=6).value)
        
        output_sheet.cell(row=row_idx, column=2).value = str(combined_row[0])
        output_sheet.cell(row=row_idx, column=5).value = str(combined_row[1])
        output_sheet.cell(row=row_idx, column=8).value = str(combined_row[2])
    
    # Save the workbook
    wb.save(input_file)

# Example usage
input_file = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\repeatability.xlsx'
combine_scores(input_file)
elapsed_time = time.time() - start_time
print(f"Total time taken: {elapsed_time/60} minutes")

def check_values_in_range(values, min_val, max_val):
    return all(min_val <= val <= max_val for val in values)

def calculate_repeatability_score(values):
    int_values = [int(value) for value in values]  # Convert string values to integers
    return 1.97 * np.std(int_values)

# Load the workbook and select the active worksheet
workbook = openpyxl.load_workbook(input_file)
sheet = workbook['Sheet1']

# Loop through each row starting from row 4
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=2, max_col=8, values_only=False):
    # Convert the string representation of lists back to lists
    col1_values = ast.literal_eval(row[0].value)  # Column B
    col4_values = ast.literal_eval(row[3].value)  # Column E
    col7_values = ast.literal_eval(row[6].value)  # Column H
    
    # Column 1, check range and calculate repeatability score
    col1_in_range = check_values_in_range(col1_values, 7, 10)
    sheet.cell(row=row[0].row, column=3).value = col1_in_range  # Column C
    sheet.cell(row=row[0].row, column=4).value = calculate_repeatability_score(col1_values)  # Column D
    
    # Column 4, check range and calculate repeatability score
    col4_in_range = check_values_in_range(col4_values, 4, 7)
    sheet.cell(row=row[0].row, column=6).value = col4_in_range  # Column F
    sheet.cell(row=row[0].row, column=7).value = calculate_repeatability_score(col4_values)  # Column G
    
    # Column 7, check range and calculate repeatability score
    col7_in_range = check_values_in_range(col7_values, 0, 4)
    sheet.cell(row=row[0].row, column=9).value = col7_in_range  # Column I
    sheet.cell(row=row[0].row, column=10).value = calculate_repeatability_score(col7_values)  # Column J

# Save the workbook
workbook.save(input_file)
