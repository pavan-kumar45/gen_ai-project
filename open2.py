import pandas as pd
from openpyxl import load_workbook
import time
start_time=time.time()
import gspread
from google.oauth2.service_account import Credentials
from google.api_core.exceptions import ResourceExhausted, InvalidArgument
import re
from langchain_openai import AzureChatOpenAI
from langchain_core.messages import HumanMessage
from config import (
    Run_specific_row,RUN_SPECIFIC_ROW_INDEX,ACCURACY_TEST,REPEATABILITY_TEST,run_column,GEMINI_TEST,OPENAI_TEST,Run_All_Rows
)
from prompts_openai import theory_prompt_template, coding_prompt_template
# Initialize the Azure OpenAI model using LangChain's AzureChatOpenAI
llm = AzureChatOpenAI(
    azure_deployment="HireGloo-OpenAI",
    api_version="2024-05-01-preview",
    temperature=0,
    max_tokens=100,
    timeout=80,
    max_retries=2,
)

# Function to generate feedback with retry mechanism using AzureChatOpenAI
def generate_feedback(prompt, retries=2, delay=1, max_wait_time=300):
    total_wait_time = 0
    for attempt in range(retries):
        if total_wait_time >= max_wait_time:
            print("Total wait time exceeded. Stopping retries.")
            return None

        try:
            system_message = (
                "system",
                "You are AI evaluator that generates score(out of 10) first and then feedback in 3 lines for user answers based on their questions. Give more weightage to concept understanding"
            )
            message = HumanMessage(content=prompt)
            response = llm.invoke([system_message, message])
            return response.content
        except (ResourceExhausted, InvalidArgument) as e:
            if isinstance(e, InvalidArgument):
                print("Invalid request. Please check your parameters.")
                return None
            if attempt < retries - 1:
                backoff_time = delay * (2 ** attempt)  # Exponential backoff
                total_wait_time += backoff_time
                print(f"Quota exceeded. Retrying in {backoff_time} seconds...")
                time.sleep(backoff_time)
            else:
                print("Max retries exceeded. Please check your quota.")
                return None
        except Exception as e:
            if "RateLimitError" in str(e):
                backoff_time = 60  # Wait for 9 seconds as suggested by the error message
                print(f"Rate limit exceeded. Retrying in {backoff_time} seconds...")
                time.sleep(backoff_time)
                continue  # Retry the request after waiting
            print(f"An unexpected error occurred: {e}")
            return None


# Function to extract score from feedback
def extract_score(feedback):
    match = re.search(r'\b\d{1,2}\b', feedback)
    return int(match.group()) if match else None

# File paths and settings for reading and writing Excel files
file_path = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\Question Bank.xlsx'
sheet_name = 'Sheet1'
column_pairs = [(0, 1), (0, 2), (0, 3)]
# column_pairs = [(0, 2), (0, 3)]


output_file_path = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\accuracy_report.xlsx'
output_sheet_name = 'Sheet2'

# Open the output workbook and select the desired sheet
wb = load_workbook(output_file_path)
ws = wb[output_sheet_name]

# Function to process all rows in the Excel sheet
def run_all_rows(flag1):
    if not flag1:
        return
    
    for col_index, col_pair in enumerate(column_pairs):
        df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=list(col_pair))
        print(f"\nReading columns {col_pair[0]+1} and {col_pair[1]+1}:\n")

        # Determine the starting columns for scores and feedback in the output file
        score_col = 2 * (col_index+1) + 2+6
        feedback_col = score_col + 1

        # Iterate through each row of the dataframe
        for index, row in df.iterrows():
            question = row.iloc[0]
            answer = row.iloc[1]

            # Determine the type of question and choose the appropriate template
            if "program" in question.lower():
                prompt_template = coding_prompt_template
            else:
                prompt_template = theory_prompt_template

            prompt = prompt_template.format(question=question, user_answer=answer)

            feedback = generate_feedback(prompt)
            if feedback:
                score = extract_score(feedback)

                # Update the output Excel file directly in the specified sheet
                ws.cell(row=index + 4, column=score_col, value=score)  # Assuming row 4 for scores
                ws.cell(row=index + 4, column=feedback_col, value=feedback)  # Assuming row 4 for feedback

                print(f"Question: {question}")
                print(f"Score: {score}")
                print(f"Feedback: {feedback}\n")

                print("-----------------------------------------------------------------------------------")
                
                # Save the workbook after each update to reflect changes immediately
                wb.save(output_file_path)

# Run the processing function with a flag to start the process
def run_specific_row(row_index, col_pair_index, flag2):
    if not flag2:
        return
    
    col_pair = column_pairs[col_pair_index]
    df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=list(col_pair))
    print(f"\nReading columns {col_pair[0]+1} and {col_pair[1]+1} for row {row_index+4}:\n")

    # Determine the starting columns for scores and feedback in the output file
    score_col = 2 * col_pair_index + 2+6
    feedback_col = score_col + 1

    # Get the specific row
    row = df.iloc[row_index]

    question = row.iloc[0]
    answer = row.iloc[1]

    # Determine the type of question and choose the appropriate template
    if "program" in question.lower():
        prompt_template = coding_prompt_template
    else:
        prompt_template = theory_prompt_template

    prompt = prompt_template.format(question=question, user_answer=answer)

    feedback = generate_feedback(prompt)
    if feedback:
        score = extract_score(feedback)

        # Update the output Excel file directly in the specified sheet
        ws.cell(row=row_index + 4, column=score_col, value=score)  # Assuming row 4 for scores
        ws.cell(row=row_index + 4, column=feedback_col, value=feedback)  # Assuming row 4 for feedback

        print(f"Question: {question}")
        print(f"Score: {score}")
        print(f"Feedback: {feedback}\n")

        print("-----------------------------------------------------------------------------------")
        
        # Save the workbook after each update to reflect changes immediately
        wb.save(output_file_path)



run_all_rows(Run_All_Rows)
run_specific_row(RUN_SPECIFIC_ROW_INDEX-4,run_column,Run_specific_row)

elapsed_time = time.time() - start_time
print(f"Total time taken: {elapsed_time/3600} hours")