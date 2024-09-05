import time
import openpyxl
from openpyxl import load_workbook
# Remove the Google-specific imports since they're no longer needed
# from google.api_core.exceptions import ResourceExhausted, InvalidArgument
from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage
from config import primary_skills, secondary_skills
from prompts_generate import coding_prompt, theory_prompt, design_prompt

# Initialize the ChatGroq model with the specified parameters
llm = ChatGroq(
    temperature=0.3,
    groq_api_key="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",  # Replace with your actual Groq API key
    model_name="llama-3.1-70b-versatile"
)

def generate_feedback(prompt, system_instruction=["Your job is to just generate questions"], retries=6, delay=1, max_wait_time=300):
    total_wait_time = 0
    for attempt in range(retries):
        if total_wait_time >= max_wait_time:
            print("Total wait time exceeded. Stopping retries.")
            return None

        try:
            # Constructing the message format as required by the Groq model
            message = HumanMessage(content=prompt, system_instruction=system_instruction)
            response = llm.invoke(message.content)
            return response.content
        except Exception as e:  # Use a generic exception handler
            print(f"Exception occurred: {e}")
            if attempt < retries - 1:
                backoff_time = delay * (2 ** attempt)  # Exponential backoff
                total_wait_time += backoff_time
                print(f"Retrying in {backoff_time} seconds...")
                time.sleep(backoff_time)
            else:
                print("Max retries exceeded. Please check your settings.")
                return None

# Function to generate questions
def generate_questions(prompt):
    response = generate_feedback(prompt)
    print(response)
    if response:
        # Assuming each line contains a question followed by its type in parentheses
        question_lines = response.split("\n")
        questions_and_types = []
        for line in question_lines:
            if '(' in line and ')' in line:
                question, q_type = map(str.strip, line.rsplit('(', 1))
                q_type = q_type.strip(')')
                questions_and_types.append((question, q_type))
        return questions_and_types
    else:
        return []

# Generate the questions
coding_questions_and_types = generate_questions(coding_prompt)
theory_questions_and_types = generate_questions(theory_prompt)
design_questions_and_types = generate_questions(design_prompt)

# Combine all questions and types into one list
all_questions_and_types = coding_questions_and_types + theory_questions_and_types + design_questions_and_types

# Path to the existing Excel file
excel_file_path = r"C:\Users\pavan\Desktop\projects\llama_3.1\CSV Files\Questions.xlsx"

# Load the workbook and select the first worksheet
wb = load_workbook(excel_file_path)
ws = wb.active

# Add headers
ws.cell(row=1, column=1, value="Questions")
ws.cell(row=1, column=2, value="Type-Skill-Difficulty")

# Add the questions and their categories to the worksheet
for i, (question, q_type) in enumerate(all_questions_and_types, start=2):
    ws.cell(row=i, column=1, value=question)
    ws.cell(row=i, column=2, value=q_type)

# Save the workbook
wb.save(excel_file_path)

print(f"Questions and categories added to {excel_file_path}")
