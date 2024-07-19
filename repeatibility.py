import pandas as pd
import time
import re
from google.api_core.exceptions import ResourceExhausted, InvalidArgument
from langchain_google_vertexai import ChatVertexAI
from langchain_core.messages import HumanMessage
import vertexai
from config import (
    SERVICE_ACCOUNT_FILE_PATH, PROJECT_ID, LOCATION,
    MODEL_PARAMS
)
from prompts import theory_prompt_template, coding_prompt_template
from openpyxl import load_workbook

# Set up the Vertex AI environment
vertexai.init(project=PROJECT_ID, location=LOCATION)

# Initialize the Vertex AI model using LangChain's ChatVertexAI
llm = ChatVertexAI(**MODEL_PARAMS)

def generate_feedback(prompt, retries=6, delay=1, max_wait_time=300):
    total_wait_time = 0
    for attempt in range(retries):
        if total_wait_time >= max_wait_time:
            print("Total wait time exceeded. Stopping retries.")
            return None

        try:
            message = HumanMessage(content=prompt)
            response = llm.invoke([message])
            return response.content
        except (ResourceExhausted, InvalidArgument) as e:
            if isinstance(e, InvalidArgument) and "topK" in str(e):
                print("Invalid topK value. Please update the value to be within the supported range.")
                return None
            if attempt < retries - 1:
                backoff_time = delay * (2 ** attempt)  # Exponential backoff
                total_wait_time += backoff_time
                print(f"Quota exceeded. Retrying in {backoff_time} seconds...")
                time.sleep(backoff_time)
            else:
                print("Max retries exceeded. Please check your quota.")
                return None

def extract_score(feedback):
    # Placeholder function to extract score from feedback
    # Replace this with the actual implementation
    score_match = re.search(r'\bscore\s*:\s*(\d+)\b', feedback, re.IGNORECASE)
    if score_match:
        return int(score_match.group(1))
    return None

def process_data(file_path, sheet_name, column_pairs, output_file_path):
    # Load the workbook
    wb = load_workbook(output_file_path)

    # Iterate over each column pair
    for col_index, col_pair in enumerate(column_pairs):
        df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=list(col_pair))
        print(f"\nReading columns {col_pair[0]+1} and {col_pair[1]+1}:\n")

        # Determine the starting columns for scores and feedback in the output file
        score_col = 2 * col_index + 2
        feedback_col = score_col + 1

        # List of sheets to update
        sheets = ['Sheet2', 'Sheet3', 'Sheet4', 'Sheet5','Sheet6',]

        # Iterate through each row of the datafram
        for index, row in df.iterrows():
            question = row.iloc[0]
            answer = row.iloc[1]

            # Determine the type of question and choose the appropriate template
            if "program" in question.lower():
                prompt_template = coding_prompt_template
            else:
                prompt_template = theory_prompt_template

            prompt = prompt_template.format(question=question, user_answer=answer)

            for iteration in range(5):  # Run each row 5 times
                feedback = generate_feedback(prompt)
                if feedback:
                    score = extract_score(feedback)

                    # Update the respective sheet for the current iteration
                    sheet = sheets[iteration]  # Select the sheet for the current iteration
                    ws = wb[sheet]
                    ws.cell(row=index + 4, column=score_col, value=score)  # Assuming row 4 for scores
                    ws.cell(row=index + 4, column=feedback_col, value=feedback)  # Assuming row 4 for feedback

                    print(f"Question: {question}")
                    print(f"Iteration: {iteration + 1}")
                    print(f"Sheet: {sheet}")
                    print(f"Score: {score}")
                    print(f"Feedback: {feedback}\n")

                    print("-----------------------------------------------------------------------------------")

                    # Save the workbook after each iteration
                    wb.save(output_file_path)

# Example usage:
file_path = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\Question Bank.xlsx'
sheet_name = 'Sheet1'
column_pairs = [(0, 1), (0, 2), (0, 3)]
output_file_path = r'C:\Users\pavan\Desktop\projects\gen ai\model-experiments\CSV FILES\repeatability.xlsx'

process_data(file_path, sheet_name, column_pairs, output_file_path)
